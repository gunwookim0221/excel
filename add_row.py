import argparse
from openpyxl import load_workbook


def find_row_indices(sheet, target_name, label_col=1):
    """Return list of row indices where cell in label_col equals target_name."""
    indices = []
    for row in range(1, sheet.max_row + 1):
        cell_val = sheet.cell(row=row, column=label_col).value
        if cell_val == target_name:
            indices.append(row)
    return indices


def pair_totals_with_excludes(total_rows, exclude_rows):
    """Pair each total row with the nearest exclude row above it."""
    pairs = []
    ex_iter = iter(exclude_rows)
    current_ex = next(ex_iter, None)
    exclude_stack = []
    for t in total_rows:
        # gather all exclude rows up to total row
        while current_ex is not None and current_ex < t:
            exclude_stack.append(current_ex)
            current_ex = next(ex_iter, None)
        if exclude_stack:
            ex_row = exclude_stack[-1]
            pairs.append((t, ex_row))
    return pairs


def insert_adjusted_rows(sheet, pairs, total_label, exclude_label, label_col=1, label_template=None):
    """Insert rows with adjusted totals using the given label template."""
    if label_template is None:
        label_template = "{total} (without {exclude})"

    offset = 0
    for total_row, exclude_row in pairs:
        tr = total_row + offset
        ex = exclude_row + offset
        sheet.insert_rows(tr + 1)
        label_cell = sheet.cell(row=tr + 1, column=label_col)
        label = label_template.format(total=total_label, exclude=exclude_label)
        label_cell.value = label
        for col in range(label_col + 1, sheet.max_column + 1):
            total_cell = sheet.cell(row=tr, column=col)
            exclude_cell = sheet.cell(row=ex, column=col)
            formula = f"={total_cell.coordinate}-{exclude_cell.coordinate}"
            sheet.cell(row=tr + 1, column=col, value=formula)
        offset += 1


def main(path, sheet_name, total_label, exclude_label, label_template=None):
    wb = load_workbook(filename=path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    ws = wb[sheet_name]

    total_rows = find_row_indices(ws, total_label)
    exclude_rows = find_row_indices(ws, exclude_label)
    if not total_rows:
        raise ValueError(f"No rows found with label '{total_label}'")
    if not exclude_rows:
        raise ValueError(f"No rows found with label '{exclude_label}'")

    pairs = pair_totals_with_excludes(total_rows, exclude_rows)
    if not pairs:
        raise ValueError("Could not pair total rows with exclude rows")

    insert_adjusted_rows(ws, pairs, total_label, exclude_label, label_template=label_template)
    wb.save(path)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Add adjusted total rows excluding a specific model")
    parser.add_argument("excel_file", help="Path to Excel workbook")
    parser.add_argument("sheet_name", help="Sheet to modify")
    parser.add_argument("total_row_name", help="Label of total rows to adjust")
    parser.add_argument("exclude_row_name", help="Label of row to exclude from totals")
    parser.add_argument("--label", dest="label_template", default=None,
                        help="Custom label for the inserted rows. Use {total} and {exclude} to reference the input labels")

    args = parser.parse_args()
    main(args.excel_file, args.sheet_name, args.total_row_name, args.exclude_row_name,
         label_template=args.label_template)
